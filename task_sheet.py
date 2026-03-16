# -*- coding: utf-8 -*-
import csv
import json
import os
import shutil
import subprocess
import tempfile
from datetime import datetime

from .setup import P, logger
from .model import ModelSheetRule, ModelChannel, ModelGroupOrder, ModelGroupProfile
from .task_base import TaskBase


class TaskSheet(TaskBase):
    @staticmethod
    def _setting(key, default=''):
        try:
            value = P.ModelSetting.get(key)
            return value if value is not None else default
        except Exception:
            return default

    @staticmethod
    def _is_true(value):
        return str(value or '').strip().lower() in ['true', 'on', '1', 'yes', 'y']

    @staticmethod
    def _run_rclone(cmd, timeout=120):
        preview = ' '.join(cmd[:12])
        logger.info(f'[ff_tvh_m3u] rclone run: {preview} ...')
        try:
            result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=timeout, text=True, encoding='utf-8', errors='replace')
        except subprocess.TimeoutExpired as e:
            return False, f'timeout: {str(e)}'
        except Exception as e:
            return False, str(e)

        if result.returncode != 0:
            return False, result.stderr or result.stdout or f'rclone failed rc={result.returncode}'
        return True, result.stdout or ''

    @staticmethod
    def _guess_kind(name='', mime_type=''):
        name = str(name or '').strip().lower()
        mime_type = str(mime_type or '').strip().lower()
        if mime_type == 'application/vnd.google-apps.spreadsheet':
            return 'google_sheet'
        if name.endswith('.xlsx') or 'spreadsheetml' in mime_type:
            return 'xlsx'
        if name.endswith('.csv') or mime_type in ['text/csv', 'application/csv']:
            return 'csv'
        return 'google_sheet'

    @staticmethod
    def list_rclone_remotes():
        cmd = ['/usr/bin/rclone', '--config', '/data/db/rclone.conf', 'listremotes']
        ok, out = TaskSheet._run_rclone(cmd, timeout=20)
        if not ok:
            return {'ret': 'danger', 'msg': f'rclone remote 목록 조회 실패: {out}', 'remotes': []}
        remotes = []
        for line in (out or '').splitlines():
            name = line.strip().rstrip(':')
            if not name:
                continue
            remotes.append(name)
        return {'ret': 'success', 'msg': f'알클론 드라이브 {len(remotes)}개 확인', 'remotes': remotes}

    @staticmethod
    def _join_remote_path(remote, path):
        remote = str(remote or '').strip().rstrip(':')
        path = str(path or '').strip().strip('/')
        if path:
            return f'{remote}:{path}'
        return f'{remote}:'

    @staticmethod
    def browse_rclone_path(remote=None, path=None):
        remote = str(remote or TaskSheet._setting('basic_sheet_rclone_remote', '')).strip().rstrip(':')
        path = str(path if path is not None else TaskSheet._setting('basic_sheet_browser_path', '')).strip().strip('/')
        if not remote:
            return {'ret': 'warning', 'msg': 'rclone remote를 먼저 선택하세요.', 'items': [], 'current_path': path}

        target = TaskSheet._join_remote_path(remote, path)
        cmd = ['/usr/bin/rclone', '--config', '/data/db/rclone.conf', 'lsjson', target, '--max-depth', '1']
        ok, out = TaskSheet._run_rclone(cmd, timeout=60)
        if not ok:
            return {'ret': 'danger', 'msg': f'rclone 경로 조회 실패: {out}', 'items': [], 'current_path': path}

        try:
            raw_items = json.loads(out or '[]')
        except Exception as e:
            return {'ret': 'danger', 'msg': f'lsjson 파싱 실패: {str(e)}', 'items': [], 'current_path': path}

        items = []
        for item in raw_items:
            if not isinstance(item, dict):
                continue
            name = str(item.get('Name') or '').strip()
            if not name:
                continue
            is_dir = bool(item.get('IsDir', False))
            item_path = str(item.get('Path') or name).strip().strip('/')
            mime_type = str(item.get('MimeType') or '').strip()
            file_id = str(item.get('ID') or '').strip()
            kind = 'folder' if is_dir else TaskSheet._guess_kind(name, mime_type)
            items.append({
                'name': name,
                'path': item_path,
                'is_dir': is_dir,
                'file_id': file_id,
                'mime_type': mime_type,
                'kind': kind,
            })

        items.sort(key=lambda x: (not x['is_dir'], x['name'].lower()))
        return {
            'ret': 'success',
            'msg': f'경로 조회 성공: {remote}:{path or "/"}',
            'items': items,
            'current_path': path,
        }

    @staticmethod
    def get_sheet_source_info():
        remote = str(TaskSheet._setting('basic_sheet_rclone_remote', '')).strip()
        file_id = str(TaskSheet._setting('basic_sheet_file_id', '')).strip()
        kind = str(TaskSheet._setting('basic_sheet_file_kind', 'google_sheet')).strip()
        selected_name = str(TaskSheet._setting('basic_sheet_selected_name', '')).strip()
        source_text = f'{remote}:{file_id}' if remote and file_id else '-'
        return {
            'ret': 'success',
            'remote': remote,
            'file_id': file_id,
            'file_kind': kind,
            'selected_name': selected_name,
            'source_text': source_text,
        }

    @staticmethod
    def _download_by_file_id():
        remote = str(TaskSheet._setting('basic_sheet_rclone_remote', '')).strip().rstrip(':')
        file_id = str(TaskSheet._setting('basic_sheet_file_id', '')).strip()
        kind = str(TaskSheet._setting('basic_sheet_file_kind', 'google_sheet')).strip()
        if not remote or not file_id:
            raise RuntimeError('rclone remote 또는 file id 가 비어 있습니다.')

        tmpdir = tempfile.mkdtemp(prefix='ff_tvh_m3u_sheet_')
        errors = []
        try:
            if kind == 'google_sheet':
                dst = os.path.join(tmpdir, 'sheet.csv')
                cmd = [
                    '/usr/bin/rclone', '--config', '/data/db/rclone.conf',
                    'backend', 'copyid', f'{remote}:', file_id, dst,
                    '--drive-export-formats', 'csv',
                ]
                ok, out = TaskSheet._run_rclone(cmd, timeout=120)
                if not ok:
                    errors.append(f'copyid_google_sheet: {out}')
                    cmd2 = cmd + ['--drive-shared-with-me']
                    ok2, out2 = TaskSheet._run_rclone(cmd2, timeout=120)
                    if not ok2:
                        errors.append(f'copyid_google_sheet_shared: {out2}')
                        raise RuntimeError(' / '.join(errors[:3]))
                return tmpdir, dst, 'csv'

            if kind == 'xlsx':
                dst = os.path.join(tmpdir, 'sheet.xlsx')
            elif kind == 'csv':
                dst = os.path.join(tmpdir, 'sheet.csv')
            else:
                raise RuntimeError(f'지원하지 않는 파일 종류: {kind}')

            cmd = ['/usr/bin/rclone', '--config', '/data/db/rclone.conf', 'backend', 'copyid', f'{remote}:', file_id, dst]
            ok, out = TaskSheet._run_rclone(cmd, timeout=120)
            if not ok:
                errors.append(f'copyid_{kind}: {out}')
                cmd2 = cmd + ['--drive-shared-with-me']
                ok2, out2 = TaskSheet._run_rclone(cmd2, timeout=120)
                if not ok2:
                    errors.append(f'copyid_{kind}_shared: {out2}')
                    raise RuntimeError(' / '.join(errors[:3]))
            return tmpdir, dst, kind
        except Exception:
            shutil.rmtree(tmpdir, ignore_errors=True)
            raise

    @staticmethod
    def _read_csv_rows(path):
        errors = []
        for encoding in ['utf-8-sig', 'utf-8', 'cp949', 'euc-kr']:
            try:
                with open(path, 'r', encoding=encoding, newline='') as f:
                    reader = csv.DictReader(f)
                    return list(reader)
            except Exception as e:
                errors.append(f'{encoding}:{str(e)}')
        raise RuntimeError('CSV 파싱 실패')

    @staticmethod
    def _read_xlsx_rows(path):
        try:
            from openpyxl import load_workbook
        except Exception as e:
            raise RuntimeError('XLSX 읽기는 openpyxl 이 필요합니다. 파일 종류를 google_sheet 또는 csv 로 바꾸거나, 서버에 openpyxl 설치가 필요합니다.') from e

        wb = load_workbook(path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(x).strip() if x is not None else '' for x in rows[0]]
        data = []
        for row in rows[1:]:
            item = {}
            empty = True
            for idx, header in enumerate(headers):
                if header == '':
                    continue
                val = row[idx] if idx < len(row) else ''
                if val not in [None, '']:
                    empty = False
                item[header] = '' if val is None else str(val).strip()
            if not empty:
                data.append(item)
        return data

    @staticmethod
    def _pick(row, keys):
        for key in keys:
            if key in row and str(row.get(key) or '').strip():
                return str(row.get(key) or '').strip()
        return ''

    @staticmethod
    def parse_sheet_rows(rows):
        items = []
        for row in rows or []:
            if not isinstance(row, dict):
                continue
            channel_name = TaskSheet._pick(row, ['이름', '채널명', 'name', 'Name'])
            aka_name = TaskSheet._pick(row, ['AKA', 'aka', '별칭', '별명'])
            group_name = TaskSheet._pick(row, ['그룹', '카테고리', 'group', 'category'])
            logo_url = TaskSheet._pick(row, ['로고', 'logo', 'icon', 'tvg-logo'])
            if group_name == '미사용':
                continue
            if not channel_name and not aka_name:
                continue
            items.append({
                'channel_name': channel_name,
                'aka_name': aka_name,
                'group_name': group_name,
                'logo_url': logo_url,
            })
        return items

    @staticmethod
    def get_sheet_rows():
        tmpdir, file_path, file_type = TaskSheet._download_by_file_id()
        try:
            if file_type == 'xlsx':
                return TaskSheet._read_xlsx_rows(file_path)
            return TaskSheet._read_csv_rows(file_path)
        finally:
            shutil.rmtree(tmpdir, ignore_errors=True)

    @staticmethod
    def import_sheet_rules():
        try:
            rows = TaskSheet.get_sheet_rows()
            items = TaskSheet.parse_sheet_rows(rows)
            if not items:
                return {'ret': 'warning', 'msg': '시트에서 읽은 규칙이 없습니다. 헤더(이름, AKA, 그룹/카테고리, 로고)를 확인하세요.'}

            ModelSheetRule.replace_all(items)
            imported_at = str(datetime.now())[:19]
            source_info = TaskSheet.get_sheet_source_info()
            P.ModelSetting.set('basic_sheet_last_import_time', imported_at)
            P.ModelSetting.set('basic_sheet_last_source', source_info.get('source_text', ''))
            logger.info(f"[ff_tvh_m3u] import_sheet_rules success rows={len(items)} source={source_info.get('source_text', '')}")
            return {
                'ret': 'success',
                'msg': f'시트 규칙 {len(items)}건을 불러왔습니다.',
                'count': len(items),
                'group_names': ModelSheetRule.get_group_names(),
                'last_import_time': imported_at,
                'source_info': source_info,
            }
        except Exception as e:
            logger.exception(f'[ff_tvh_m3u] import_sheet_rules exception: {str(e)}')
            return {'ret': 'danger', 'msg': f'구글 시트 불러오기 실패: {str(e)}'}

    @staticmethod
    def ensure_sheet_rules(use_cached_on_fail=True):
        ret = TaskSheet.import_sheet_rules()
        if ret.get('ret') == 'success':
            return ret
        cached = ModelSheetRule.get_count()
        if use_cached_on_fail and cached > 0:
            logger.warning(f"[ff_tvh_m3u] ensure_sheet_rules fallback to cached rules: {ret.get('msg')}")
            return {
                'ret': 'warning',
                'msg': f'시트 새로고침은 실패했지만 기존 캐시 규칙 {cached}건을 사용합니다.',
                'count': cached,
                'group_names': ModelSheetRule.get_group_names(),
            }
        return ret

    @staticmethod
    def match_sheet_channels():
        try:
            if not TaskSheet._is_true(TaskSheet._setting('basic_channel_auto_match', 'False')):
                return {'ret': 'warning', 'msg': '채널 자동분배를 먼저 켜세요.'}

            channels = ModelChannel.get_all()
            if not channels:
                return {'ret': 'warning', 'msg': '먼저 채널 동기화를 실행하세요.'}

            refresh_ret = TaskSheet.ensure_sheet_rules(use_cached_on_fail=True)
            if refresh_ret.get('ret') not in ['success', 'warning']:
                return refresh_ret

            updates = []
            matched_count = 0
            for row in channels:
                rule = ModelSheetRule.find_match(row.name)
                if rule is None:
                    continue
                matched_count += 1
                updates.append({
                    'channel_uuid': row.channel_uuid,
                    'sheet_group_name': str(rule.group_name or '').strip(),
                    'sheet_logo_url': str(rule.logo_url or '').strip(),
                })

            ModelChannel.replace_sheet_matches(updates)
            ModelGroupOrder.sync_from_group_names(ModelChannel.get_effective_group_names())
            ModelGroupProfile.cleanup_by_group_names(ModelChannel.get_effective_group_names())

            matched_at = str(datetime.now())[:19]
            P.ModelSetting.set('basic_sheet_last_match_count', str(matched_count))

            suffix = ''
            if refresh_ret.get('ret') == 'warning':
                suffix = ' / 시트 새로고침 실패로 기존 캐시 규칙 사용'

            logger.info(f'[ff_tvh_m3u] match_sheet_channels done matched={matched_count} channels={len(channels)}')
            return {
                'ret': 'success',
                'msg': f'채널 매칭 완료: {matched_count}개 채널 매칭 / 규칙 {ModelSheetRule.get_count()}개{suffix}',
                'matched_count': matched_count,
                'channel_count': len(channels),
                'rule_count': ModelSheetRule.get_count(),
                'matched_at': matched_at,
            }
        except Exception as e:
            logger.exception(f'[ff_tvh_m3u] match_sheet_channels exception: {str(e)}')
            return {'ret': 'danger', 'msg': f'채널 매칭 실패: {str(e)}'}


    @staticmethod
    def clear_sheet_matches(reason='manual'):
        try:
            cleared_count = ModelChannel.clear_sheet_matches()
            group_names = ModelChannel.get_effective_group_names()
            ModelGroupOrder.sync_from_group_names(group_names)
            ModelGroupProfile.cleanup_by_group_names(group_names)
            P.ModelSetting.set('basic_sheet_last_match_count', '0')

            if reason == 'auto_match_off':
                msg = '자동 채널 매칭을 꺼서 기존 시트 매칭 결과를 초기화했습니다.'
            else:
                msg = f'기존 시트 매칭 결과 {cleared_count}건을 초기화했습니다.'

            logger.info(f'[ff_tvh_m3u] clear_sheet_matches done cleared={cleared_count} reason={reason}')
            return {
                'ret': 'success',
                'msg': msg,
                'cleared_count': cleared_count,
            }
        except Exception as e:
            logger.exception(f'[ff_tvh_m3u] clear_sheet_matches exception: {str(e)}')
            return {'ret': 'danger', 'msg': f'시트 매칭 초기화 실패: {str(e)}'}

    @staticmethod
    def get_sheet_group_names():
        return ModelSheetRule.get_group_names()
